"""Signal list processing."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, BinaryIO, Protocol

from openpyxl import load_workbook

REQUIRED_COLUMNS = ["IOA", "Bezeichnung", "Einheit"]


class SupportsUploadFile(Protocol):
    filename: str
    file: BinaryIO


class SignalListError(Exception):
    """Raised when a signal list cannot be processed."""


def read_signal_list(file: SupportsUploadFile, destination: Path) -> dict[str, Any]:
    """Read an uploaded Excel file and store a normalized JSON representation."""

    destination.parent.mkdir(parents=True, exist_ok=True)
    data = file.file.read()
    destination.write_bytes(data)

    workbook = load_workbook(destination, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(max_row=1), [])
    headers = [cell.value for cell in header_row]
    if not headers:
        raise SignalListError("Empty spreadsheet")

    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise SignalListError(f"Missing required columns: {', '.join(missing)}")

    column_indices = {name: headers.index(name) for name in REQUIRED_COLUMNS}
    records = []
    for row in sheet.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue
        try:
            ioa = int(row[column_indices["IOA"]].value)
        except (TypeError, ValueError) as exc:
            raise SignalListError("IOA column must contain integers") from exc
        label = str(row[column_indices["Bezeichnung"]].value or "")
        unit = str(row[column_indices["Einheit"]].value or "")
        records.append({"ioa": ioa, "label": label, "unit": unit})

    json_destination = destination.with_suffix(".json")
    json_destination.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"records": records, "path": json_destination.name}


__all__ = ["read_signal_list", "SignalListError", "REQUIRED_COLUMNS", "SupportsUploadFile"]
